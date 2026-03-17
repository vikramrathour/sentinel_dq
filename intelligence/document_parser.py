"""
DocumentParser — Phase 2 OrianDQ Intelligence Module.

Detects the type of a structured reference document (JSON Schema, XSD,
dbt YAML schema, or generic YAML data contract) and extracts raw
constraints from it.  The output is consumed by RuleExtractor to produce
typed DQRule dicts.
"""

from __future__ import annotations

import io
import json
import logging
from typing import Any

import yaml

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_XSD_NS = {
    "xs": "http://www.w3.org/2001/XMLSchema",
    "xsd": "http://www.w3.org/2001/XMLSchema",
}

# JSON Schema formats that are worth surfacing as explicit rules.
_KNOWN_FORMATS = {"email", "uri", "date", "date-time", "uuid", "hostname", "ipv4", "ipv6"}

# XSD built-in types → normalised name used in raw constraints.
_XSD_TYPE_MAP = {
    "xs:string": "string",
    "xsd:string": "string",
    "xs:integer": "integer",
    "xsd:integer": "integer",
    "xs:int": "integer",
    "xsd:int": "integer",
    "xs:long": "integer",
    "xsd:long": "integer",
    "xs:short": "integer",
    "xsd:short": "integer",
    "xs:decimal": "decimal",
    "xsd:decimal": "decimal",
    "xs:float": "float",
    "xsd:float": "float",
    "xs:double": "float",
    "xsd:double": "float",
    "xs:date": "date",
    "xsd:date": "date",
    "xs:dateTime": "date-time",
    "xsd:dateTime": "date-time",
    "xs:boolean": "boolean",
    "xsd:boolean": "boolean",
}


# ---------------------------------------------------------------------------
# Main class
# ---------------------------------------------------------------------------


class DocumentParser:
    """
    Parses structured reference documents and returns raw constraints.

    Supported document types
    ------------------------
    - JSON Schema  (.json with ``$schema`` or ``type/properties`` keys)
    - XSD          (.xsd or XML containing ``xs:schema`` / ``xsd:schema``)
    - dbt YAML     (.yaml/.yml with ``version:`` + ``models:``/``sources:``)
    - YAML contract(.yaml/.yml with ``dataset:`` or ``columns:`` at root)
    """

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self, file_bytes: bytes, filename: str) -> dict:
        """
        Auto-detect document type and extract raw constraints.

        Parameters
        ----------
        file_bytes:
            Raw bytes of the document.
        filename:
            Original filename (used for extension-based detection).

        Returns
        -------
        dict with keys:
            ``document_type``   – one of "json_schema" | "xsd" | "yaml_dbt" |
                                  "yaml_contract" | "unknown"
            ``detected_columns``– list of column/field names found
            ``raw_constraints`` – list of raw constraint dicts (pre-rule)
            ``metadata``        – document-level metadata dict
        """
        empty: dict = {
            "document_type": "unknown",
            "detected_columns": [],
            "raw_constraints": [],
            "metadata": {},
        }

        if not file_bytes:
            logger.warning("document_parser: received empty file_bytes for '%s'", filename)
            return empty

        doc_type = self._detect_type(file_bytes, filename)
        logger.debug("document_parser: detected type '%s' for file '%s'", doc_type, filename)

        try:
            if doc_type == "json_schema":
                data = json.loads(file_bytes.decode("utf-8", errors="replace"))
                result = self._parse_json_schema(data)
            elif doc_type == "xsd":
                result = self._parse_xsd(file_bytes)
            elif doc_type == "yaml_dbt":
                data = yaml.safe_load(file_bytes.decode("utf-8", errors="replace"))
                result = self._parse_yaml_dbt(data)
            elif doc_type == "yaml_contract":
                data = yaml.safe_load(file_bytes.decode("utf-8", errors="replace"))
                result = self._parse_yaml_contract(data)
            elif doc_type == "docx_dictionary":
                result = self._parse_docx_dictionary(file_bytes)
            elif doc_type == "excel_catalog":
                result = self._parse_excel_catalog(file_bytes)
            else:
                return empty
        except Exception as exc:  # noqa: BLE001
            logger.error("document_parser: failed to parse '%s': %s", filename, exc)
            return {**empty, "document_type": doc_type}

        result["document_type"] = doc_type
        return result

    # ------------------------------------------------------------------
    # Detection
    # ------------------------------------------------------------------

    def _detect_type(self, file_bytes: bytes, filename: str) -> str:
        """
        Heuristically determine the document type.

        Detection order: XSD → JSON Schema → dbt YAML → YAML contract → unknown.
        """
        name_lower = filename.lower()
        ext = name_lower.rsplit(".", 1)[-1] if "." in name_lower else ""
        preview = file_bytes[:2000]

        # ---- XSD -------------------------------------------------------
        if ext == "xsd":
            return "xsd"
        if b"xs:schema" in preview or b"xsd:schema" in preview:
            return "xsd"

        # ---- JSON Schema -----------------------------------------------
        if ext == "json":
            try:
                data = json.loads(file_bytes.decode("utf-8", errors="replace"))
                if isinstance(data, dict):
                    has_schema_key = "$schema" in data
                    has_object_shape = (
                        data.get("type") == "object" and "properties" in data
                    )
                    if has_schema_key or has_object_shape:
                        return "json_schema"
            except (json.JSONDecodeError, UnicodeDecodeError):
                pass

        # ---- YAML variants ---------------------------------------------
        if ext in ("yaml", "yml"):
            try:
                text = file_bytes.decode("utf-8", errors="replace")
                data = yaml.safe_load(text)
                if isinstance(data, dict):
                    has_version = "version" in data
                    # dbt: version + models or sources
                    if has_version and ("models" in data or "sources" in data):
                        return "yaml_dbt"
                    # Generic contract: dataset or columns at root
                    if "dataset" in data or "columns" in data:
                        return "yaml_contract"
            except yaml.YAMLError:
                pass

        # ---- DOCX data dictionary --------------------------------------
        if ext == "docx":
            return "docx_dictionary"

        # ---- Excel catalog ---------------------------------------------
        if ext in ("xlsx", "xls"):
            return "excel_catalog"

        return "unknown"

    # ------------------------------------------------------------------
    # JSON Schema parser
    # ------------------------------------------------------------------

    def _parse_json_schema(self, data: dict) -> dict:
        """
        Extract raw constraints from a JSON Schema document.

        Handles: required, type, minimum, maximum, minLength, maxLength,
        pattern, enum, format, nullable.
        """
        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        if not isinstance(data, dict):
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        # Top-level metadata
        for meta_key in ("title", "description", "$schema", "id", "$id"):
            if meta_key in data:
                metadata[meta_key] = data[meta_key]

        properties: dict = data.get("properties") or {}
        required_fields: list = data.get("required") or []

        # Collect columns
        detected_columns = list(properties.keys())

        # Required at schema level
        for field in required_fields:
            raw_constraints.append(
                {
                    "field": field,
                    "constraint": "required",
                    "value": True,
                    "source_path": f"$.required['{field}']",
                }
            )

        # Per-property constraints
        for field, prop_def in properties.items():
            if not isinstance(prop_def, dict):
                continue

            base_path = f"$.properties.{field}"

            # nullable: false  (OpenAPI extension, also honoured in JSON Schema)
            if prop_def.get("nullable") is False:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "required",
                        "value": True,
                        "source_path": f"{base_path}.nullable",
                    }
                )

            # type
            type_val = prop_def.get("type")
            if type_val:
                # type can be a list e.g. ["string", "null"]
                if isinstance(type_val, list):
                    non_null_types = [t for t in type_val if t != "null"]
                    type_val = non_null_types[0] if non_null_types else None
                if type_val:
                    raw_constraints.append(
                        {
                            "field": field,
                            "constraint": "type",
                            "value": type_val,
                            "source_path": f"{base_path}.type",
                        }
                    )

            # minimum / maximum
            for kw in ("minimum", "exclusiveMinimum"):
                if kw in prop_def:
                    raw_constraints.append(
                        {
                            "field": field,
                            "constraint": "minimum",
                            "value": prop_def[kw],
                            "source_path": f"{base_path}.{kw}",
                        }
                    )
            for kw in ("maximum", "exclusiveMaximum"):
                if kw in prop_def:
                    raw_constraints.append(
                        {
                            "field": field,
                            "constraint": "maximum",
                            "value": prop_def[kw],
                            "source_path": f"{base_path}.{kw}",
                        }
                    )

            # minLength / maxLength
            if "minLength" in prop_def:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "minLength",
                        "value": prop_def["minLength"],
                        "source_path": f"{base_path}.minLength",
                    }
                )
            if "maxLength" in prop_def:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "maxLength",
                        "value": prop_def["maxLength"],
                        "source_path": f"{base_path}.maxLength",
                    }
                )

            # pattern
            if "pattern" in prop_def:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "pattern",
                        "value": prop_def["pattern"],
                        "source_path": f"{base_path}.pattern",
                    }
                )

            # enum
            if "enum" in prop_def:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "enum",
                        "value": prop_def["enum"],
                        "source_path": f"{base_path}.enum",
                    }
                )

            # format (only well-known formats)
            fmt = prop_def.get("format")
            if fmt and fmt in _KNOWN_FORMATS:
                raw_constraints.append(
                    {
                        "field": field,
                        "constraint": "format",
                        "value": fmt,
                        "source_path": f"{base_path}.format",
                    }
                )

        return {
            "detected_columns": detected_columns,
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    # ------------------------------------------------------------------
    # XSD parser
    # ------------------------------------------------------------------

    def _parse_xsd(self, file_bytes: bytes) -> dict:
        """
        Extract raw constraints from an XML Schema Definition document.

        Uses lxml.etree.  Walks ALL xs:element nodes in the schema (not
        just top-level) and extracts type, occurrence, and restriction
        constraints.
        """
        try:
            from lxml import etree  # local import — optional dependency
        except ImportError:
            logger.error("document_parser: lxml is required to parse XSD files")
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        try:
            root = etree.fromstring(file_bytes)
        except etree.XMLSyntaxError as exc:
            logger.error("document_parser: XSD parse error: %s", exc)
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        # Resolve the active namespace prefix (xs or xsd)
        ns_uri = "http://www.w3.org/2001/XMLSchema"
        ns = {"xs": ns_uri}

        # Schema-level metadata
        target_ns = root.get("targetNamespace")
        if target_ns:
            metadata["targetNamespace"] = target_ns

        # Walk every element declaration in the document
        for elem in root.iter(f"{{{ns_uri}}}element"):
            name = elem.get("name")
            if not name:
                continue

            detected_columns.append(name)
            base_xpath = f"/xs:schema//xs:element[@name='{name}']"

            # Occurrence / required
            min_occurs = elem.get("minOccurs", "1")  # default is "1"
            nillable = elem.get("nillable", "false").lower()

            if nillable == "true":
                # nillable means xsi:nil="true" is allowed — mark as nullable
                raw_constraints.append(
                    {
                        "field": name,
                        "constraint": "nullable",
                        "value": True,
                        "source_path": f"{base_xpath}/@nillable",
                    }
                )
            elif min_occurs != "0":
                # minOccurs not "0" means the element is required
                raw_constraints.append(
                    {
                        "field": name,
                        "constraint": "required",
                        "value": True,
                        "source_path": f"{base_xpath}/@minOccurs",
                    }
                )

            # Inline type attribute
            type_attr = elem.get("type")
            if type_attr:
                normalised = _XSD_TYPE_MAP.get(type_attr, type_attr.split(":")[-1])
                raw_constraints.append(
                    {
                        "field": name,
                        "constraint": "type",
                        "value": normalised,
                        "source_path": f"{base_xpath}/@type",
                    }
                )

            # xs:simpleType / xs:complexType → xs:restriction children
            self._extract_xsd_restrictions(elem, name, base_xpath, raw_constraints, ns_uri)

        return {
            "detected_columns": detected_columns,
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    def _extract_xsd_restrictions(
        self,
        element_node: Any,
        field_name: str,
        base_xpath: str,
        raw_constraints: list[dict],
        ns_uri: str,
    ) -> None:
        """
        Recursively find xs:restriction under an xs:element and extract
        facet constraints (pattern, enum, range, length, type).
        """
        tag = lambda local: f"{{{ns_uri}}}{local}"  # noqa: E731

        for restriction in element_node.iter(tag("restriction")):
            base = restriction.get("base", "")
            if base:
                normalised_base = _XSD_TYPE_MAP.get(base, base.split(":")[-1])
                raw_constraints.append(
                    {
                        "field": field_name,
                        "constraint": "type",
                        "value": normalised_base,
                        "source_path": f"{base_xpath}/xs:restriction/@base",
                    }
                )

            # Collect enumeration values together
            enum_values: list[str] = []
            for child in restriction:
                local = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                value = child.get("value")
                if value is None:
                    continue

                if local == "enumeration":
                    enum_values.append(value)
                elif local in ("minInclusive", "minExclusive"):
                    raw_constraints.append(
                        {
                            "field": field_name,
                            "constraint": "minimum",
                            "value": self._coerce_number(value),
                            "source_path": f"{base_xpath}/xs:restriction/xs:{local}",
                        }
                    )
                elif local in ("maxInclusive", "maxExclusive"):
                    raw_constraints.append(
                        {
                            "field": field_name,
                            "constraint": "maximum",
                            "value": self._coerce_number(value),
                            "source_path": f"{base_xpath}/xs:restriction/xs:{local}",
                        }
                    )
                elif local == "pattern":
                    raw_constraints.append(
                        {
                            "field": field_name,
                            "constraint": "pattern",
                            "value": value,
                            "source_path": f"{base_xpath}/xs:restriction/xs:pattern",
                        }
                    )
                elif local == "maxLength":
                    raw_constraints.append(
                        {
                            "field": field_name,
                            "constraint": "maxLength",
                            "value": int(value),
                            "source_path": f"{base_xpath}/xs:restriction/xs:maxLength",
                        }
                    )
                elif local == "minLength":
                    raw_constraints.append(
                        {
                            "field": field_name,
                            "constraint": "minLength",
                            "value": int(value),
                            "source_path": f"{base_xpath}/xs:restriction/xs:minLength",
                        }
                    )

            if enum_values:
                raw_constraints.append(
                    {
                        "field": field_name,
                        "constraint": "enum",
                        "value": enum_values,
                        "source_path": f"{base_xpath}/xs:restriction/xs:enumeration",
                    }
                )

    # ------------------------------------------------------------------
    # dbt YAML parser
    # ------------------------------------------------------------------

    def _parse_yaml_dbt(self, data: dict) -> dict:
        """
        Extract raw constraints from a dbt ``schema.yml`` file.

        Handles models and sources blocks.  Supported tests: not_null,
        unique, accepted_values, relationships, dbt_utils.accepted_range.
        """
        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        if not isinstance(data, dict):
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        metadata["version"] = data.get("version")

        # Both models and sources share the same column-test structure.
        model_lists: list[list] = []
        if "models" in data and isinstance(data["models"], list):
            model_lists.append(data["models"])
        if "sources" in data and isinstance(data["sources"], list):
            # sources have a tables sub-key
            for source in data["sources"]:
                if isinstance(source, dict) and isinstance(source.get("tables"), list):
                    model_lists.append(source["tables"])
                    metadata.setdefault("source_name", source.get("name"))

        for model_list in model_lists:
            for model in model_list:
                if not isinstance(model, dict):
                    continue

                model_name = model.get("name", "")
                metadata.setdefault("model_name", model_name)

                columns: list = model.get("columns") or []
                for col in columns:
                    if not isinstance(col, dict):
                        continue

                    col_name = col.get("name")
                    if not col_name:
                        continue

                    detected_columns.append(col_name)
                    base_path = f"models['{model_name}'].columns['{col_name}']"

                    if col.get("description"):
                        metadata.setdefault("column_descriptions", {})[col_name] = col[
                            "description"
                        ]

                    tests: list = col.get("tests") or []
                    for test in tests:
                        self._parse_dbt_test(
                            test, col_name, base_path, raw_constraints
                        )

        return {
            "detected_columns": list(dict.fromkeys(detected_columns)),  # preserve order, dedupe
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    def _parse_dbt_test(
        self,
        test: Any,
        col_name: str,
        base_path: str,
        raw_constraints: list[dict],
    ) -> None:
        """
        Translate a single dbt test entry into a raw constraint dict.

        A test can be a plain string (e.g. ``not_null``) or a dict with a
        single key that maps to parameters (e.g. ``{accepted_values: {values: []}}``.
        """
        if isinstance(test, str):
            test_name = test
            test_params: dict = {}
        elif isinstance(test, dict):
            # Typically only one key
            test_name = next(iter(test), "")
            test_params = test.get(test_name) or {}
            if not isinstance(test_params, dict):
                test_params = {}
        else:
            return

        src_path = f"{base_path}.tests.{test_name}"

        if test_name == "not_null":
            raw_constraints.append(
                {
                    "field": col_name,
                    "constraint": "required",
                    "value": True,
                    "source_path": src_path,
                }
            )
        elif test_name == "unique":
            raw_constraints.append(
                {
                    "field": col_name,
                    "constraint": "unique",
                    "value": True,
                    "source_path": src_path,
                }
            )
        elif test_name == "accepted_values":
            values = test_params.get("values")
            if values is not None:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "enum",
                        "value": list(values),
                        "source_path": src_path,
                    }
                )
        elif test_name in ("relationships", "dbt_utils.relationships"):
            # Record referential integrity as a custom constraint — do not fail.
            raw_constraints.append(
                {
                    "field": col_name,
                    "constraint": "custom",
                    "value": {"test": "relationships", "params": test_params},
                    "source_path": src_path,
                }
            )
        elif test_name in ("dbt_utils.accepted_range", "accepted_range"):
            constraint: dict = {}
            if "min_value" in test_params:
                constraint["minimum"] = test_params["min_value"]
            if "max_value" in test_params:
                constraint["maximum"] = test_params["max_value"]
            if constraint:
                # Emit separate min/max so RuleExtractor can merge them.
                if "minimum" in constraint:
                    raw_constraints.append(
                        {
                            "field": col_name,
                            "constraint": "minimum",
                            "value": constraint["minimum"],
                            "source_path": f"{src_path}.min_value",
                        }
                    )
                if "maximum" in constraint:
                    raw_constraints.append(
                        {
                            "field": col_name,
                            "constraint": "maximum",
                            "value": constraint["maximum"],
                            "source_path": f"{src_path}.max_value",
                        }
                    )
        else:
            # Unknown test — record as custom so nothing is silently lost.
            raw_constraints.append(
                {
                    "field": col_name,
                    "constraint": "custom",
                    "value": {"test": test_name, "params": test_params},
                    "source_path": src_path,
                }
            )

    # ------------------------------------------------------------------
    # Generic YAML contract parser
    # ------------------------------------------------------------------

    def _parse_yaml_contract(self, data: dict) -> dict:
        """
        Extract raw constraints from a generic YAML data contract.

        Supports: nullable, allowed_values (enum), min/max (range),
        pattern, min_length/max_length, type.  Unknown keys are skipped.
        """
        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        if not isinstance(data, dict):
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        # Top-level metadata
        for key in ("dataset", "version", "description", "owner", "domain"):
            if key in data:
                metadata[key] = data[key]

        columns: list = data.get("columns") or []
        if not isinstance(columns, list):
            columns = []

        for col in columns:
            if not isinstance(col, dict):
                continue

            col_name = col.get("name")
            if not col_name:
                continue

            detected_columns.append(col_name)
            base_path = f"columns['{col_name}']"

            if col.get("description"):
                metadata.setdefault("column_descriptions", {})[col_name] = col[
                    "description"
                ]

            # type
            col_type = col.get("type")
            if col_type:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "type",
                        "value": str(col_type).lower(),
                        "source_path": f"{base_path}.type",
                    }
                )

            # nullable: false → required
            nullable = col.get("nullable", True)
            if nullable is False or str(nullable).lower() == "false":
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "required",
                        "value": True,
                        "source_path": f"{base_path}.nullable",
                    }
                )

            # allowed_values → enum
            allowed = col.get("allowed_values")
            if allowed is not None and isinstance(allowed, (list, tuple)):
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "enum",
                        "value": list(allowed),
                        "source_path": f"{base_path}.allowed_values",
                    }
                )

            # min / max → range
            if "min" in col:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "minimum",
                        "value": col["min"],
                        "source_path": f"{base_path}.min",
                    }
                )
            if "max" in col:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "maximum",
                        "value": col["max"],
                        "source_path": f"{base_path}.max",
                    }
                )

            # pattern
            if "pattern" in col:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "pattern",
                        "value": col["pattern"],
                        "source_path": f"{base_path}.pattern",
                    }
                )

            # min_length / max_length
            if "min_length" in col:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "minLength",
                        "value": col["min_length"],
                        "source_path": f"{base_path}.min_length",
                    }
                )
            if "max_length" in col:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "maxLength",
                        "value": col["max_length"],
                        "source_path": f"{base_path}.max_length",
                    }
                )

            # format (if explicitly specified)
            fmt = col.get("format")
            if fmt and fmt in _KNOWN_FORMATS:
                raw_constraints.append(
                    {
                        "field": col_name,
                        "constraint": "format",
                        "value": fmt,
                        "source_path": f"{base_path}.format",
                    }
                )

        return {
            "detected_columns": detected_columns,
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    # ------------------------------------------------------------------
    # DOCX data-dictionary parser
    # ------------------------------------------------------------------

    def _parse_docx_dictionary(self, file_bytes: bytes) -> dict:
        """
        Extract raw constraints from a Word (.docx) data dictionary.

        Each table in the document is scanned.  The first row is treated
        as a header; subsequent rows are data rows.  Header columns are
        mapped to semantic roles via ``_detect_header_columns()``.
        """
        try:
            from docx import Document  # python-docx
        except ImportError:
            logger.error("document_parser: python-docx is required to parse DOCX files")
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        doc = Document(io.BytesIO(file_bytes))

        # Document-level metadata
        try:
            title = doc.core_properties.title
            if title:
                metadata["title"] = title
        except Exception:  # noqa: BLE001
            pass

        tables_found = 0

        for table_idx, table in enumerate(doc.tables):
            rows = table.rows
            if len(rows) < 2:
                # Need at least a header row + one data row
                continue
            if len(table.columns) < 2:
                continue

            tables_found += 1

            # Build header list (normalised)
            header_cells = rows[0].cells
            headers: list[str] = [c.text.lower().strip() for c in header_cells]

            col_map = self._detect_header_columns(headers)
            if "name" not in col_map:
                # Cannot identify the field-name column — skip this table
                continue

            for row_idx, row in enumerate(rows[1:], start=1):
                cells = row.cells
                cell_texts = [c.text.strip() for c in cells]

                # Field name
                name_idx = col_map["name"]
                if name_idx >= len(cell_texts):
                    continue
                field_name = cell_texts[name_idx].strip()
                if not field_name:
                    continue

                if field_name not in detected_columns:
                    detected_columns.append(field_name)

                # Nullable / Required
                if "nullable" in col_map:
                    nul_idx = col_map["nullable"]
                    if nul_idx < len(cell_texts):
                        nul_val = cell_texts[nul_idx].lower().strip()
                        not_nullable_tokens = {
                            "no", "n", "false", "required", "mandatory",
                            "not null", "yes*",
                        }
                        if nul_val in not_nullable_tokens:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "required",
                                "value": True,
                                "source_path": (
                                    f"table[{table_idx}].row[{row_idx}].nullable"
                                ),
                            })

                # Type
                if "type" in col_map:
                    type_idx = col_map["type"]
                    if type_idx < len(cell_texts):
                        type_val = cell_texts[type_idx].lower().strip()
                        if type_val:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "type",
                                "value": type_val,
                                "source_path": (
                                    f"table[{table_idx}].row[{row_idx}].type"
                                ),
                            })

                # Enum / valid values
                if "enum" in col_map:
                    enum_idx = col_map["enum"]
                    if enum_idx < len(cell_texts):
                        enum_raw = cell_texts[enum_idx].strip()
                        if enum_raw:
                            # Try to detect inline range patterns first
                            range_handled = False

                            # "Min: 18, Max: 99" pattern
                            import re
                            min_max_pat = re.search(
                                r"min[:\s]+(\d+(?:\.\d+)?)[,;\s]+max[:\s]+(\d+(?:\.\d+)?)",
                                enum_raw,
                                re.IGNORECASE,
                            )
                            if min_max_pat:
                                raw_constraints.append({
                                    "field": field_name,
                                    "constraint": "minimum",
                                    "value": self._coerce_number(min_max_pat.group(1)),
                                    "source_path": (
                                        f"table[{table_idx}].row[{row_idx}].enum"
                                    ),
                                })
                                raw_constraints.append({
                                    "field": field_name,
                                    "constraint": "maximum",
                                    "value": self._coerce_number(min_max_pat.group(2)),
                                    "source_path": (
                                        f"table[{table_idx}].row[{row_idx}].enum"
                                    ),
                                })
                                range_handled = True

                            # "18-99" bare range pattern
                            range_pat = re.fullmatch(
                                r"(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)",
                                enum_raw,
                            )
                            if range_pat and not range_handled:
                                raw_constraints.append({
                                    "field": field_name,
                                    "constraint": "minimum",
                                    "value": self._coerce_number(range_pat.group(1)),
                                    "source_path": (
                                        f"table[{table_idx}].row[{row_idx}].enum"
                                    ),
                                })
                                raw_constraints.append({
                                    "field": field_name,
                                    "constraint": "maximum",
                                    "value": self._coerce_number(range_pat.group(2)),
                                    "source_path": (
                                        f"table[{table_idx}].row[{row_idx}].enum"
                                    ),
                                })
                                range_handled = True

                            if not range_handled:
                                # Split by comma or newline
                                sep = "," if "," in enum_raw else "\n"
                                values = [
                                    v.strip() for v in enum_raw.split(sep) if v.strip()
                                ]
                                if values:
                                    raw_constraints.append({
                                        "field": field_name,
                                        "constraint": "enum",
                                        "value": values,
                                        "source_path": (
                                            f"table[{table_idx}].row[{row_idx}].enum"
                                        ),
                                    })

                # Min
                if "min" in col_map:
                    min_idx = col_map["min"]
                    if min_idx < len(cell_texts):
                        min_raw = cell_texts[min_idx].strip()
                        if min_raw:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "minimum",
                                "value": self._coerce_number(min_raw),
                                "source_path": (
                                    f"table[{table_idx}].row[{row_idx}].min"
                                ),
                            })

                # Max
                if "max" in col_map:
                    max_idx = col_map["max"]
                    if max_idx < len(cell_texts):
                        max_raw = cell_texts[max_idx].strip()
                        if max_raw:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "maximum",
                                "value": self._coerce_number(max_raw),
                                "source_path": (
                                    f"table[{table_idx}].row[{row_idx}].max"
                                ),
                            })

                # Pattern / format / regex
                if "pattern" in col_map:
                    pat_idx = col_map["pattern"]
                    if pat_idx < len(cell_texts):
                        pat_raw = cell_texts[pat_idx].strip()
                        if pat_raw:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "pattern",
                                "value": pat_raw,
                                "source_path": (
                                    f"table[{table_idx}].row[{row_idx}].pattern"
                                ),
                            })

        metadata["tables_found"] = tables_found

        return {
            "detected_columns": detected_columns,
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    # ------------------------------------------------------------------
    # Excel data-catalog parser
    # ------------------------------------------------------------------

    def _parse_excel_catalog(self, file_bytes: bytes) -> dict:
        """
        Extract raw constraints from an Excel (.xlsx/.xls) data catalog.

        Finds the most likely sheet (by name keywords), locates the header
        row, and processes data rows the same way as the DOCX parser.
        """
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:
            logger.error("document_parser: openpyxl is required to parse Excel files")
            return {"detected_columns": [], "raw_constraints": [], "metadata": {}}

        raw_constraints: list[dict] = []
        detected_columns: list[str] = []
        metadata: dict = {}

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        sheet_names = wb.sheetnames
        metadata["total_sheets"] = len(sheet_names)

        # Pick the best sheet by name keywords
        catalog_keywords = {
            "dict", "catalog", "schema", "field", "column",
            "metadata", "definition",
        }
        selected_sheet = sheet_names[0]
        for sname in sheet_names:
            if any(kw in sname.lower() for kw in catalog_keywords):
                selected_sheet = sname
                break

        metadata["sheet_used"] = selected_sheet
        ws = wb[selected_sheet]

        # Collect all rows as lists of string values
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            str_row = [
                str(cell).strip() if cell is not None else ""
                for cell in row
            ]
            all_rows.append(str_row)

        if not all_rows:
            return {
                "detected_columns": detected_columns,
                "raw_constraints": raw_constraints,
                "metadata": metadata,
            }

        # Find the header row: first row where ≥2 cells match role keywords
        header_row_idx: int | None = None
        col_map: dict = {}

        for i, row in enumerate(all_rows):
            candidate_map = self._detect_header_columns(
                [c.lower() for c in row]
            )
            if len(candidate_map) >= 2:
                header_row_idx = i
                col_map = candidate_map
                break

        if header_row_idx is None or "name" not in col_map:
            # No usable header found
            return {
                "detected_columns": detected_columns,
                "raw_constraints": raw_constraints,
                "metadata": metadata,
            }

        data_rows = all_rows[header_row_idx + 1:]

        consecutive_empty = 0
        import re  # noqa: PLC0415 — already imported above in docx path

        for row_idx, row in enumerate(data_rows):
            # Skip / detect empty rows
            if all(c == "" for c in row):
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    break
                continue
            consecutive_empty = 0

            # Field name
            name_idx = col_map.get("name")
            if name_idx is None or name_idx >= len(row):
                continue
            field_name = row[name_idx].strip()
            if not field_name:
                continue

            if field_name not in detected_columns:
                detected_columns.append(field_name)

            src_prefix = f"sheet[{selected_sheet}].row[{row_idx}]"

            # Nullable / Required
            if "nullable" in col_map:
                nul_idx = col_map["nullable"]
                if nul_idx < len(row):
                    nul_val = row[nul_idx].lower().strip()
                    not_nullable_tokens = {
                        "no", "n", "false", "required", "mandatory",
                        "not null", "yes*",
                    }
                    if nul_val in not_nullable_tokens:
                        raw_constraints.append({
                            "field": field_name,
                            "constraint": "required",
                            "value": True,
                            "source_path": f"{src_prefix}.nullable",
                        })

            # Type
            if "type" in col_map:
                type_idx = col_map["type"]
                if type_idx < len(row):
                    type_val = row[type_idx].lower().strip()
                    if type_val:
                        raw_constraints.append({
                            "field": field_name,
                            "constraint": "type",
                            "value": type_val,
                            "source_path": f"{src_prefix}.type",
                        })

            # Enum / valid values
            if "enum" in col_map:
                enum_idx = col_map["enum"]
                if enum_idx < len(row):
                    enum_raw = row[enum_idx].strip()
                    if enum_raw:
                        range_handled = False

                        min_max_pat = re.search(
                            r"min[:\s]+(\d+(?:\.\d+)?)[,;\s]+max[:\s]+(\d+(?:\.\d+)?)",
                            enum_raw,
                            re.IGNORECASE,
                        )
                        if min_max_pat:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "minimum",
                                "value": self._coerce_number(min_max_pat.group(1)),
                                "source_path": f"{src_prefix}.enum",
                            })
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "maximum",
                                "value": self._coerce_number(min_max_pat.group(2)),
                                "source_path": f"{src_prefix}.enum",
                            })
                            range_handled = True

                        range_pat = re.fullmatch(
                            r"(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)",
                            enum_raw,
                        )
                        if range_pat and not range_handled:
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "minimum",
                                "value": self._coerce_number(range_pat.group(1)),
                                "source_path": f"{src_prefix}.enum",
                            })
                            raw_constraints.append({
                                "field": field_name,
                                "constraint": "maximum",
                                "value": self._coerce_number(range_pat.group(2)),
                                "source_path": f"{src_prefix}.enum",
                            })
                            range_handled = True

                        if not range_handled:
                            sep = "," if "," in enum_raw else "\n"
                            values = [
                                v.strip() for v in enum_raw.split(sep) if v.strip()
                            ]
                            if values:
                                raw_constraints.append({
                                    "field": field_name,
                                    "constraint": "enum",
                                    "value": values,
                                    "source_path": f"{src_prefix}.enum",
                                })

            # Min
            if "min" in col_map:
                min_idx = col_map["min"]
                if min_idx < len(row):
                    min_raw = row[min_idx].strip()
                    if min_raw:
                        raw_constraints.append({
                            "field": field_name,
                            "constraint": "minimum",
                            "value": self._coerce_number(min_raw),
                            "source_path": f"{src_prefix}.min",
                        })

            # Max
            if "max" in col_map:
                max_idx = col_map["max"]
                if max_idx < len(row):
                    max_raw = row[max_idx].strip()
                    if max_raw:
                        raw_constraints.append({
                            "field": field_name,
                            "constraint": "maximum",
                            "value": self._coerce_number(max_raw),
                            "source_path": f"{src_prefix}.max",
                        })

            # Pattern
            if "pattern" in col_map:
                pat_idx = col_map["pattern"]
                if pat_idx < len(row):
                    pat_raw = row[pat_idx].strip()
                    if pat_raw:
                        raw_constraints.append({
                            "field": field_name,
                            "constraint": "pattern",
                            "value": pat_raw,
                            "source_path": f"{src_prefix}.pattern",
                        })

        return {
            "detected_columns": detected_columns,
            "raw_constraints": raw_constraints,
            "metadata": metadata,
        }

    # ------------------------------------------------------------------
    # Header-column role detector  (shared by DOCX and Excel parsers)
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_header_columns(headers: list[str]) -> dict:
        """
        Map normalised header strings to semantic role names.

        Parameters
        ----------
        headers:
            List of already-lowercased, stripped header strings.

        Returns
        -------
        dict mapping role name → column index, e.g.
        ``{"name": 0, "type": 2, "nullable": 3}``.
        Roles not detected are omitted.
        """
        col_map: dict[str, int] = {}

        for idx, h in enumerate(headers):
            h = h.lower().strip()

            # Field name — must not be a description column
            if "description" not in h and "desc" not in h:
                if any(kw in h for kw in ("field", "column", "name", "attribute", "element")):
                    col_map.setdefault("name", idx)
                    continue

            # Data type
            if any(kw in h for kw in ("type", "datatype", "data type")):
                col_map.setdefault("type", idx)
                continue

            # Nullable / required
            if any(
                kw in h
                for kw in ("nullable", "required", "mandatory", "null", "optional")
            ):
                col_map.setdefault("nullable", idx)
                continue

            # Valid values / enum
            if any(
                kw in h
                for kw in ("valid value", "allowed value", "value", "enum", "domain")
            ):
                col_map.setdefault("enum", idx)
                continue

            # Min
            if h in ("min", "minimum") or h.startswith("min"):
                col_map.setdefault("min", idx)
                continue

            # Max
            if h in ("max", "maximum") or h.startswith("max"):
                col_map.setdefault("max", idx)
                continue

            # Pattern / format / regex
            if any(kw in h for kw in ("pattern", "format", "regex", "mask")):
                col_map.setdefault("pattern", idx)
                continue

            # Description (last — catch-all for remaining desc-like headers)
            if any(kw in h for kw in ("description", "desc", "comment", "remarks")):
                col_map.setdefault("description", idx)
                continue

        return col_map

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    @staticmethod
    def _coerce_number(value: str) -> int | float:
        """Try to parse a string as int, then float, else return the string."""
        try:
            int_val = int(value)
            return int_val
        except (ValueError, TypeError):
            pass
        try:
            return float(value)
        except (ValueError, TypeError):
            return value  # type: ignore[return-value]
